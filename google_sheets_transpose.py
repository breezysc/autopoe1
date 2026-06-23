import pandas as pd
import openpyxl
from openpyxl.styles import Font
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox, StringVar
import os
import re

class GoogleSheetsTransposeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Google Sheets Transpose Tool")
        self.root.geometry("550x450")
        
        # Variables
        self.sheets_url = StringVar(value="https://docs.google.com/spreadsheets/d/1b6L5DxAQOAj7jftwHyMc7WOZON-CvnnlLdL3x_yd3Uw/edit#gid=408339131")
        self.sheet_name = StringVar(value="April 6")
        self.range_input = StringVar(value="Q3:Q19")
        self.local_file = StringVar()
        
        self.create_widgets()
    
    def create_widgets(self):
        # UI Layout in English
        Label(self.root, text="Google Sheets URL:").grid(row=0, column=0, padx=10, pady=20, sticky="e")
        Entry(self.root, textvariable=self.sheets_url, width=45).grid(row=0, column=1, columnspan=2)

        Label(self.root, text="Date / Name:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.sheet_name, width=30).grid(row=1, column=1, sticky="w")
        
        Label(self.root, text="Data Range:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.range_input, width=30).grid(row=2, column=1, sticky="w")
        
        Label(self.root, text="Local Template:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.local_file, width=30).grid(row=3, column=1, sticky="w")
        Button(self.root, text="Browse", command=self.browse_file).grid(row=3, column=2)
        
        # The START button
        Button(self.root, text="START", command=self.execute_transpose, width=25, bg="#0078D7", fg="white", font=("Arial", 10, "bold")).grid(row=5, column=1, pady=40)

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename: self.local_file.set(filename)

    def col_to_index(self, col_str):
        index = 0
        for char in col_str.upper():
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index

    def parse_range(self, range_str):
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str.upper())
        return match.groups() if match else None

    def is_merged_slave(self, ws, cell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                if cell.coordinate != merged_range.start_cell.coordinate:
                    return True
        return False

    def execute_transpose(self):
        url = self.sheets_url.get()
        range_str = self.range_input.get()
        template_path = self.local_file.get()
        name_val = self.sheet_name.get()
        
        if not all([url, range_str, template_path]):
            messagebox.showerror("Error", "Please fill in all fields.")
            return

        try:
            # 1. Read Data
            gid_match = re.search(r'gid=(\d+)', url)
            gid = gid_match.group(1) if gid_match else "0"
            base_url = url.split('/edit')[0]
            csv_url = f"{base_url}/export?format=csv&gid={gid}"
            df = pd.read_csv(csv_url, header=None)
            
            # 2. Extract Data
            start_col_let, start_row, end_col_let, end_row = self.parse_range(range_str)
            col_idx = self.col_to_index(start_col_let) - 1
            data_slice = df.iloc[int(start_row)-1 : int(end_row), col_idx].fillna("").tolist()
            
            # 3. Process Excel
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Find next available row in Column B
            target_row = 1
            while True:
                cell_b = ws.cell(row=target_row, column=2)
                if self.is_merged_slave(ws, cell_b) or cell_b.value is not None:
                    target_row += 1
                    continue
                break
            
            # Write Name to Column A
            cell_a = ws.cell(row=target_row, column=1)
            if not self.is_merged_slave(ws, cell_a):
                cell_a.value = name_val
            
            # Write Data starting from Column B
            current_col_ptr = 2 
            last_written_cell = None
            
            for val in data_slice:
                while True:
                    cell = ws.cell(row=target_row, column=current_col_ptr)
                    if self.is_merged_slave(ws, cell):
                        current_col_ptr += 1
                        continue
                    cell.value = val
                    last_written_cell = cell
                    current_col_ptr += 1
                    break
            
            # Bold last cell
            if last_written_cell:
                last_written_cell.font = Font(bold=True)
            
            # 6. Save
            new_path = os.path.join(os.path.dirname(template_path), f"{name_val}.xlsx")
            wb.save(new_path)
            
            # Done
            messagebox.showinfo("Success", f"Process completed!\nFile saved as: {name_val}.xlsx")

        except Exception as e:
            messagebox.showerror("Failed", f"Error: {str(e)}")

if __name__ == "__main__":
    root = Tk()
    app = GoogleSheetsTransposeApp(root)
    root.mainloop()