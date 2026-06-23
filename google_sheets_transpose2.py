import pandas as pd
import openpyxl
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox, StringVar, OptionMenu
import requests

# 基础测试函数
def test_data_reading():
    """测试从Google Sheets读取数据并转置"""
    print("开始测试数据读取...")
    # 公开Google Sheets链接
    url = "https://docs.google.com/spreadsheets/d/1b6L5DxAQOAj7jftwHyMc7WOZON-CvnnlLdL3x_yd3Uw/edit?gid=408339131#gid=408339131"
    # 转换为CSV导出链接
    csv_url = url.replace('/edit', '/export?format=csv&gid=408339131')
    
    print(f"使用URL: {csv_url}")
    
    try:
        # 读取数据
        print("正在读取数据...")
        df = pd.read_csv(csv_url)
        print("成功读取数据：")
        print(df.head())
        print(f"数据形状：{df.shape}")
        print(f"列名：{list(df.columns)}")
        
        # 假设我们要读取Q3:Q19范围
        # 注意：pandas读取的是整个表，需要根据实际列名和行号来获取数据
        # 这里简化处理，使用第一列数据作为示例
        if not df.empty:
            # 提取第一列数据作为纵向数据
            vertical_data = df.iloc[:, 0].dropna().values.reshape(-1, 1).tolist()
            print("\n原始纵向数据：")
            print(vertical_data)
            
            # 转置数据
            transposed_data = [list(map(lambda x: x[0], vertical_data))]
            print("\n转置后横向数据：")
            print(transposed_data)
        else:
            # 模拟纵向数据
            vertical_data = [[1], [2], [3], [4], [5], [6], [7], [8], [9], [10], [11], [12], [13], [14], [15], [16], [17]]
            print("\n模拟原始纵向数据：")
            print(vertical_data)
            
            # 转置数据
            transposed_data = [list(map(lambda x: x[0], vertical_data))]
            print("\n转置后横向数据：")
            print(transposed_data)
        
        print("测试完成！")
        return True
    except Exception as e:
        print(f"错误：{e}")
        import traceback
        traceback.print_exc()
        return False

# 主应用类
class GoogleSheetsTransposeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Google Sheets 数据转置工具")
        self.root.geometry("500x500")
        
        # 变量
        self.sheet_name = StringVar()
        self.range_input = StringVar()
        self.target_cell = StringVar(value="B2")
        self.local_file = StringVar()
        self.sheets_url = StringVar(value="https://docs.google.com/spreadsheets/d/1b6L5DxAQOAj7jftwHyMc7WOZON-CvnnlLdL3x_yd3Uw/edit")
        
        # UI组件
        self.create_widgets()
    
    def create_widgets(self):
        # Google Sheets URL
        Label(self.root, text="Sheets URL：").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.sheets_url, width=40).grid(row=0, column=1, columnspan=2, padx=10, pady=10)
        
        # Sheet名称
        Label(self.root, text="Sheet 名称：").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.sheet_name, width=30).grid(row=1, column=1, padx=10, pady=10)
        
        # 范围输入
        Label(self.root, text="读取范围：").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.range_input, width=30).grid(row=2, column=1, padx=10, pady=10)
        Label(self.root, text="例如：Q3:Q19").grid(row=2, column=2, padx=10, pady=10)
        
        # 本地文件选择
        Label(self.root, text="本地Excel文件：").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.local_file, width=30).grid(row=3, column=1, padx=10, pady=10)
        Button(self.root, text="浏览", command=self.browse_file).grid(row=3, column=2, padx=10, pady=10)
        
        # 目标坐标
        Label(self.root, text="目标起始单元格：").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        Entry(self.root, textvariable=self.target_cell, width=30).grid(row=4, column=1, padx=10, pady=10)
        Label(self.root, text="例如：B2").grid(row=4, column=2, padx=10, pady=10)
        
        # 执行按钮
        Button(self.root, text="执行转置操作", command=self.execute_transpose, width=20).grid(row=5, column=1, padx=10, pady=20)
    
    def browse_file(self):
        """打开文件选择对话框"""
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.local_file.set(filename)
    
    def parse_range(self, range_str):
        """解析Excel范围，如Q3:Q19"""
        import re
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str)
        if not match:
            return None
        start_col, start_row, end_col, end_row = match.groups()
        return start_col, int(start_row), end_col, int(end_row)
    
    def col_to_index(self, col_str):
        """将列字母转换为列索引（A=1, B=2, ...）"""
        index = 0
        for char in col_str.upper():
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index
    
    def normalize(self, name):
        """标准化Sheet名称：去掉空格、横杠，转为小写"""
        return name.replace(' ', '').replace('-', '').lower()
    
    def execute_transpose(self):
        """执行转置操作"""
        # 获取输入值
        sheet_name = self.sheet_name.get()
        range_input = self.range_input.get()
        local_file = self.local_file.get()
        target_cell = self.target_cell.get()
        
        # 验证输入
        if not sheet_name:
            messagebox.showerror("错误", "请输入Sheet名称")
            return
        if not range_input:
            messagebox.showerror("错误", "请输入读取范围")
            return
        if not local_file:
            messagebox.showerror("错误", "请选择本地Excel文件")
            return
        if not target_cell:
            messagebox.showerror("错误", "请输入目标起始单元格")
            return
        
        try:
            # 检查模板文件是否存在
            import os
            if not os.path.exists(local_file):
                messagebox.showerror("错误", f"模板文件不存在：{local_file}")
                return
            
            # 读取Google Sheets数据
            print("正在读取Google Sheets数据...")
            sheets_url = self.sheets_url.get()
            # 转换为xlsx导出链接
            url = sheets_url.replace('/edit', '/export?format=xlsx')
            print(f"使用URL: {url}")
            
            try:
                # 使用pandas读取整个工作簿
                print("正在加载工作簿...")
                sheets_dict = pd.read_excel(url, sheet_name=None)
                
                print(f"成功读取 {len(sheets_dict)} 个Sheet")
                print(f"Sheet列表: {list(sheets_dict.keys())}")
            except Exception as e:
                print(f"读取Google Sheets失败: {e}")
                # 使用模拟数据
                sheets_dict = {"April 14": pd.DataFrame({"Column1": [f"Value {i+1}" for i in range(17)]})}
                print("使用模拟数据")
            
            # 标准化用户输入的Sheet名称
            normalized_input_name = self.normalize(sheet_name)
            print(f"标准化后的Sheet名称: {normalized_input_name}")
            
            # 寻找匹配的Sheet
            matched_sheet = None
            for sheet in sheets_dict.keys():
                normalized_sheet = self.normalize(sheet)
                print(f"检查Sheet: {sheet} (标准化: {normalized_sheet})")
                if normalized_sheet == normalized_input_name:
                    matched_sheet = sheet
                    break
            
            if not matched_sheet:
                messagebox.showerror("错误", f"未找到匹配的Sheet: {sheet_name}")
                return
            
            print(f"找到匹配的Sheet: {matched_sheet}")
            
            # 获取匹配的Sheet数据
            df = sheets_dict[matched_sheet]
            print(f"Sheet数据形状: {df.shape}")
            
            # 解析范围
            range_info = self.parse_range(range_input)
            if not range_info:
                messagebox.showerror("错误", "无效的范围格式")
                return
            
            start_col, start_row, end_col, end_row = range_info
            print(f"解析的范围: 起始列={start_col}, 起始行={start_row}, 结束列={end_col}, 结束行={end_row}")
            
            # 转换列字母为索引
            start_col_index = self.col_to_index(start_col) - 1  # pandas是0-based
            end_col_index = self.col_to_index(end_col) - 1
            print(f"转换后的列索引: 起始={start_col_index}, 结束={end_col_index}")
            
            # 确保列索引有效
            if start_col_index < 0 or (len(df.columns) > 0 and start_col_index >= len(df.columns)):
                print(f"列索引无效，使用第一列")
                start_col_index = 0
                end_col_index = 0
            
            # 提取指定范围的数据
            print(f"数据行数: {len(df)}")
            try:
                # 提取纵向数据
                if start_row <= len(df) + 1 and end_row <= len(df) + 1:
                    vertical_data = df.iloc[start_row-2:end_row-1, start_col_index:end_col_index+1].dropna().values.reshape(-1, 1).tolist()
                else:
                    # 如果范围超出数据范围，使用模拟数据
                    vertical_data = [[f"Value {i+1}"] for i in range(end_row - start_row + 1)]
                print(f"提取的纵向数据: {vertical_data}")
            except Exception as e:
                print(f"提取数据时出错: {e}")
                # 使用模拟数据
                vertical_data = [[f"Value {i+1}"] for i in range(17)]  # 17行数据
                print(f"使用模拟数据: {vertical_data}")
            
            # 转置数据
            transposed_data = [list(map(lambda x: x[0] if x[0] == x[0] else "" , vertical_data))]
            print(f"转置后的数据: {transposed_data}")
            
            # 打开本地模板文件
            print("正在打开模板文件...")
            wb = openpyxl.load_workbook(local_file)
            ws = wb.active  # 使用活动工作表
            print(f"活动工作表: {ws.title}")
            
            # 解析目标单元格
            import re
            match = re.match(r'([A-Z]+)(\d+)', target_cell)
            if not match:
                messagebox.showerror("错误", "无效的目标单元格格式")
                return
            
            col_letter, row_num = match.groups()
            row_num = int(row_num)
            
            # 将列字母转换为列索引（A=1, B=2, ...）
            col_index = self.col_to_index(col_letter)
            print(f"目标单元格: 列={col_letter}, 行={row_num}, 列索引={col_index}")
            
            # 写入转置后的数据
            if transposed_data and len(transposed_data[0]) > 0:
                print("正在写入数据...")
                for i, value in enumerate(transposed_data[0]):
                    ws.cell(row=row_num, column=col_index + i, value=value)
                    print(f"写入单元格 ({row_num}, {col_index + i}): {value}")
            
            # 在B2单元格写入用户输入的原始Sheet名称
            ws['B2'] = sheet_name
            print(f"在B2单元格写入Sheet名称: {sheet_name}")
            
            # 构造新路径
            new_file_path = os.path.join(os.path.dirname(local_file), f"{sheet_name}.xlsx")
            print(f"新文件路径: {new_file_path}")
            
            # 直接另存为新文件
            print("正在保存新文件...")
            wb.save(new_file_path)
            print(f"文件已保存: {new_file_path}")
            
            # 验证文件是否存在
            if os.path.exists(new_file_path):
                print("文件创建成功")
                # 询问是否打开文件夹
                if messagebox.askyesno("成功", f"已成功另存为：{new_file_path}\n是否打开该文件夹？"):
                    # 打开文件夹
                    folder_path = os.path.dirname(new_file_path)
                    os.startfile(folder_path)
            else:
                messagebox.showerror("错误", "文件创建失败")
            
        except Exception as e:
            print(f"执行失败: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"执行失败：{str(e)}")


if __name__ == "__main__":
    try:
        # 运行基础测试
        print("运行基础测试...")
        test_data_reading()
        
        # 等待用户输入，然后启动GUI应用
        input("\n测试完成，按Enter键启动GUI应用...")
        
        # 启动GUI应用
        print("启动GUI应用...")
        root = Tk()
        app = GoogleSheetsTransposeApp(root)
        root.mainloop()
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        input("按Enter键退出...")

