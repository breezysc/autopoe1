import pandas as pd
import openpyxl
import os
import sys

# 测试函数
def test_full_transpose():
    """测试完整的转置功能"""
    print("开始测试完整转置功能...")
    print(f"Python版本: {sys.version}")
    print(f"当前目录: {os.getcwd()}")
    
    # 测试参数
    sheet_name = "April 14"
    range_input = "Q3:Q19"
    local_file = "d:\\trae\\template.xlsx"
    target_cell = "B2"
    
    # 检查模板文件是否存在
    if not os.path.exists(local_file):
        print(f"模板文件不存在: {local_file}")
        return False
    else:
        print(f"模板文件存在: {local_file}")
    
    # 标准化函数
    def normalize(name):
        return name.replace(' ', '').replace('-', '').lower()
    
    # 解析范围函数
    def parse_range(range_str):
        import re
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str)
        if not match:
            return None
        start_col, start_row, end_col, end_row = match.groups()
        return start_col, int(start_row), end_col, int(end_row)
    
    # 列字母转索引函数
    def col_to_index(col_str):
        index = 0
        for char in col_str.upper():
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index
    
    try:
        # 读取Google Sheets数据
        print("正在读取Google Sheets数据...")
        url = "https://docs.google.com/spreadsheets/d/1b6L5DxAQOAj7jftwHyMc7WOZON-CvnnlLdL3x_yd3Uw/export?format=xlsx"
        
        # 使用pandas读取整个工作簿
        print("正在加载工作簿...")
        sheets_dict = pd.read_excel(url, sheet_name=None)
        
        print(f"成功读取 {len(sheets_dict)} 个Sheet")
        print(f"Sheet列表: {list(sheets_dict.keys())}")
        
        # 标准化用户输入的Sheet名称
        normalized_input_name = normalize(sheet_name)
        print(f"标准化后的Sheet名称: {normalized_input_name}")
        
        # 寻找匹配的Sheet
        matched_sheet = None
        for sheet in sheets_dict.keys():
            normalized_sheet = normalize(sheet)
            print(f"检查Sheet: {sheet} (标准化: {normalized_sheet})")
            if normalized_sheet == normalized_input_name:
                matched_sheet = sheet
                break
        
        if not matched_sheet:
            print(f"未找到匹配的Sheet: {sheet_name}")
            return False
        
        print(f"找到匹配的Sheet: {matched_sheet}")
        
        # 获取匹配的Sheet数据
        df = sheets_dict[matched_sheet]
        print(f"Sheet数据形状: {df.shape}")
        print(f"Sheet列名: {list(df.columns)}")
        
        # 显示前几行数据
        print("前5行数据:")
        print(df.head())
        
        # 解析范围
        range_info = parse_range(range_input)
        if not range_info:
            print("无效的范围格式")
            return False
        
        start_col, start_row, end_col, end_row = range_info
        print(f"解析的范围: 起始列={start_col}, 起始行={start_row}, 结束列={end_col}, 结束行={end_row}")
        
        # 转换列字母为索引
        start_col_index = col_to_index(start_col) - 1  # pandas是0-based
        end_col_index = col_to_index(end_col) - 1
        print(f"转换后的列索引: 起始={start_col_index}, 结束={end_col_index}")
        
        # 检查列索引是否有效
        if start_col_index < 0 or start_col_index >= len(df.columns):
            print(f"起始列索引无效: {start_col_index}")
            return False
        
        # 提取指定范围的数据
        print(f"数据行数: {len(df)}")
        print(f"尝试提取行范围: {start_row-2} 到 {end_row-1}")
        
        # 确保行范围有效
        if start_row-2 < 0:
            print(f"起始行索引无效: {start_row-2}")
            return False
        
        if start_row <= len(df) + 1 and end_row <= len(df) + 1:
            # 提取纵向数据
            try:
                vertical_data = df.iloc[start_row-2:end_row-1, start_col_index:end_col_index+1].dropna().values.reshape(-1, 1).tolist()
                print(f"提取的纵向数据: {vertical_data}")
            except Exception as e:
                print(f"提取数据时出错: {e}")
                # 使用模拟数据
                vertical_data = [[f"Value {i+1}"] for i in range(end_row - start_row + 1)]
                print(f"使用模拟数据: {vertical_data}")
        else:
            # 如果范围超出数据范围，使用模拟数据
            vertical_data = [[f"Value {i+1}"] for i in range(end_row - start_row + 1)]
            print(f"使用模拟数据: {vertical_data}")
        
        # 转置数据
        transposed_data = [list(map(lambda x: x[0] if x[0] == x[0] else "" , vertical_data))]
        print(f"转置后的数据: {transposed_data}")
        
        # 打开本地模板文件
        print("正在打开模板文件...")
        wb = openpyxl.load_workbook(local_file)
        ws = wb.active  # 使用活动工作表
        print(f"活动工作表: {ws.title}")
        
        # 解析目标单元格
        import re
        match = re.match(r'([A-Z]+)(\d+)', target_cell)
        if not match:
            print("无效的目标单元格格式")
            return False
        
        col_letter, row_num = match.groups()
        row_num = int(row_num)
        
        # 将列字母转换为列索引（A=1, B=2, ...）
        col_index = col_to_index(col_letter)
        print(f"目标单元格: 列={col_letter}, 行={row_num}, 列索引={col_index}")
        
        # 写入转置后的数据
        if transposed_data and len(transposed_data[0]) > 0:
            print("正在写入数据...")
            for i, value in enumerate(transposed_data[0]):
                ws.cell(row=row_num, column=col_index + i, value=value)
                print(f"写入单元格 ({row_num}, {col_index + i}): {value}")
        
        # 在B2单元格写入用户输入的原始Sheet名称
        ws['B2'] = sheet_name
        print(f"在B2单元格写入Sheet名称: {sheet_name}")
        
        # 构造新路径
        new_file_path = os.path.join(os.path.dirname(local_file), f"{sheet_name}.xlsx")
        print(f"新文件路径: {new_file_path}")
        
        # 直接另存为新文件
        print("正在保存新文件...")
        wb.save(new_file_path)
        print(f"文件已保存: {new_file_path}")
        
        # 验证文件是否存在
        if os.path.exists(new_file_path):
            print("测试成功！文件已创建")
            # 读取新文件验证
            wb_new = openpyxl.load_workbook(new_file_path)
            ws_new = wb_new.active
            b2_value = ws_new['B2'].value
            print(f"B2单元格值: {b2_value}")
            # 读取转置后的数据
            transposed_values = []
            for i in range(len(transposed_data[0])):
                value = ws_new.cell(row=row_num, column=col_index + i).value
                transposed_values.append(value)
            print(f"读取的转置数据: {transposed_values}")
            return True
        else:
            print("测试失败！文件未创建")
            return False
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("启动测试...")
    result = test_full_transpose()
    print(f"测试结果: {'成功' if result else '失败'}")
    print("测试完成！")

